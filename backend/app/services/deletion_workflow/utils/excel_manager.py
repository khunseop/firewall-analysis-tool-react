# app/services/deletion_workflow/utils/excel_manager.py
"""
Excel 파일 관리 유틸리티.
fpat/fpat/policy_deletion_processor/utils/excel_manager.py 이식.
"""

import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

logger = logging.getLogger(__name__)


class ExcelManager:
    """Excel 파일 스타일 적용 및 저장을 담당하는 클래스"""

    def __init__(self, config_manager):
        self.config = config_manager

    def save_to_excel(self, df, sheet_type: str, file_name: str):
        """
        DataFrame을 기존 Excel 파일의 지정 시트에 저장하고 스타일을 적용합니다.

        Args:
            df: 저장할 DataFrame
            sheet_type: 대상 시트 이름
            file_name: Excel 파일 경로
        """
        try:
            wb = load_workbook(file_name)
            sheet = wb[sheet_type]

            sheet.insert_rows(1)
            sheet['A1'] = '="대상 정책 수: "&COUNTA(B:B)-1'
            sheet['A1'].font = Font(bold=True)

            header_color = self.config.get('excel_styles.header_fill_color', 'E0E0E0')
            history_color = self.config.get('excel_styles.history_fill_color', 'ccffff')

            for col in range(1, 8):
                cell = sheet.cell(row=2, column=col)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')

            if sheet_type != '이력없음_미사용정책':
                for col in range(8, 24):
                    cell = sheet.cell(row=2, column=col)
                    cell.fill = PatternFill(start_color=history_color, end_color=history_color, fill_type='solid')

            wb.save(file_name)
            logger.info(f"Excel '{file_name}' — '{sheet_type}' 시트 저장 완료")
        except Exception as e:
            logger.exception(f"Excel 저장 실패: {e}")
            raise
