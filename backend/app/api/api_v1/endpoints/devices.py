from typing import List, Any
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging

from app import crud, models, schemas
from app.db.session import get_db
from app.services import device_service

router = APIRouter()
logger = logging.getLogger(__name__)

@router.post("/", response_model=schemas.Device)
async def create_device(
    device_in: schemas.DeviceCreate,
    db: AsyncSession = Depends(get_db)
):
    if device_in.password != device_in.password_confirm:
        raise HTTPException(status_code=400, detail="Passwords do not match")
    db_device = await crud.device.get_device_by_name(db, name=device_in.name)
    if db_device:
        raise HTTPException(status_code=400, detail="Device with this name already registered")
    return await crud.device.create_device(db=db, device=device_in)

@router.get("/", response_model=List[schemas.Device])
async def read_devices(
    skip: int = 0,
    limit: int | None = None,
    db: AsyncSession = Depends(get_db)
):
    """장비 목록 조회 (limit이 None이면 모든 장비 조회)"""
    devices = await crud.device.get_devices(db, skip=skip, limit=limit)
    return devices


@router.get("/dashboard/stats", response_model=schemas.DashboardStatsResponse)
async def get_dashboard_stats(
    db: AsyncSession = Depends(get_db)
):
    """대시보드 통계 조회 (전체 통계 + 장비별 통계)"""
    return await crud.device.get_dashboard_stats(db)

@router.get("/excel-template")
async def download_excel_template():
    """엑셀 서식 파일 다운로드"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "장비 등록"
        
        # 헤더 정의
        headers = [
            "이름*",
            "IP 주소*",
            "벤더*",
            "사용자명*",
            "비밀번호*",
            "HA Peer IP",
            "설명",
            "SSH로 마지막 매칭일시 수집 (TRUE/FALSE)"
        ]
        
        # 헤더 작성
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 예시 데이터 작성
        example_data = [
            ["PaloAlto-FW-01", "192.168.1.1", "paloalto", "admin", "password123", "192.168.1.2", "본사 방화벽", "FALSE"],
            ["NGF-FW-01", "192.168.2.1", "ngf", "admin", "password123", "", "지사 방화벽", "FALSE"],
        ]
        
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        example_font = Font(size=10)
        
        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = example_fill
                cell.font = example_font
        
        # 설명 시트 추가
        ws2 = wb.create_sheet("설명")
        instructions = [
            ["필드명", "설명", "필수 여부", "예시"],
            ["이름", "장비의 고유 이름", "필수", "PaloAlto-FW-01"],
            ["IP 주소", "장비의 IP 주소", "필수", "192.168.1.1"],
            ["벤더", "벤더 코드 (paloalto, ngf, mock)", "필수", "paloalto"],
            ["사용자명", "로그인 사용자명", "필수", "admin"],
            ["비밀번호", "로그인 비밀번호", "필수", "password123"],
            ["HA Peer IP", "HA 구성 시 Peer IP (선택)", "선택", "192.168.1.2"],
            ["설명", "장비 설명 (선택)", "선택", "본사 방화벽"],
            ["SSH로 마지막 매칭일시 수집", "TRUE 또는 FALSE (기본값: FALSE)", "선택", "FALSE"],
            ["", "", "", ""],
            ["주의사항", "", "", ""],
            ["- * 표시된 필드는 필수 입력 항목입니다.", "", "", ""],
            ["- 벤더 코드는 정확히 입력해야 합니다: paloalto, ngf, mock", "", "", ""],
            ["- 이름과 IP 주소는 중복될 수 없습니다.", "", "", ""],
            ["- 예시 행은 삭제하고 실제 데이터를 입력하세요.", "", "", ""],
        ]
        
        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
        
        # 열 너비 조정
        column_widths = [20, 15, 12, 12, 15, 15, 30, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 설명 시트 열 너비 조정
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 20
        
        # 파일 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response_headers = {"Content-Disposition": 'attachment; filename="device_template.xlsx"'}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=response_headers,
        )
    except Exception as e:
        logger.error(f"Failed to generate Excel template: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"엑셀 서식 생성 실패: {str(e)}")

@router.get("/{device_id}", response_model=schemas.Device)
async def read_device(
    device_id: int,
    db: AsyncSession = Depends(get_db)
):
    db_device = await crud.device.get_device(db, device_id=device_id)
    if db_device is None:
        raise HTTPException(status_code=404, detail="Device not found")
    return db_device

@router.put("/{device_id}", response_model=schemas.Device)
async def update_device(
    device_id: int,
    device_in: schemas.DeviceUpdate,
    db: AsyncSession = Depends(get_db)
):
    if device_in.password and device_in.password != device_in.password_confirm:
        raise HTTPException(status_code=400, detail="Passwords do not match")
    db_device = await crud.device.get_device(db, device_id=device_id)
    if db_device is None:
        raise HTTPException(status_code=404, detail="Device not found")
    updated_device = await crud.device.update_device(db=db, db_obj=db_device, obj_in=device_in)
    return updated_device

@router.delete("/{device_id}", response_model=schemas.Device)
async def delete_device(
    device_id: int,
    db: AsyncSession = Depends(get_db)
):
    db_device = await crud.device.remove_device(db, id=device_id)
    if db_device is None:
        raise HTTPException(status_code=404, detail="Device not found")
    return db_device

@router.post("/{device_id}/test-connection", response_model=dict)
async def test_connection(
    device_id: int,
    db: AsyncSession = Depends(get_db)
) -> Any:
    """
    Test the connection to a device.
    """
    db_device = await crud.device.get_device(db, device_id=device_id)
    if db_device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    connection_result = await device_service.test_device_connection(db_device)

    if connection_result["status"] == "failure":
        raise HTTPException(status_code=400, detail=connection_result["message"])

    return connection_result

@router.post("/bulk-import")
async def bulk_import_devices(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """엑셀 파일을 통한 장비 일괄 등록"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.")
    
    try:
        # 파일 읽기
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # 헤더 확인 및 매핑
        header_row = [cell.value for cell in ws[1]]
        header_mapping = {
            "이름*": "name",
            "IP 주소*": "ip_address",
            "벤더*": "vendor",
            "사용자명*": "username",
            "비밀번호*": "password",
            "HA Peer IP": "ha_peer_ip",
            "설명": "description",
            "SSH로 마지막 매칭일시 수집 (TRUE/FALSE)": "use_ssh_for_last_hit_date"
        }
        
        # 필수 필드 확인
        required_fields = ["이름*", "IP 주소*", "벤더*", "사용자명*", "비밀번호*"]
        missing_fields = [field for field in required_fields if field not in header_row]
        if missing_fields:
            raise HTTPException(
                status_code=400,
                detail=f"필수 컬럼이 없습니다: {', '.join(missing_fields)}"
            )
        
        # 데이터 파싱
        devices = []
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 건너뛰기
            if not any(row):
                continue
            
            device_data = {}
            for col_idx, header in enumerate(header_row):
                if header not in header_mapping:
                    continue
                
                field_name = header_mapping[header]
                value = row[col_idx]
                
                # 문자열 값 정리
                if isinstance(value, str):
                    value = value.strip()
                    if not value:
                        value = None
                
                # 필수 필드 검증
                if header in required_fields and not value:
                    errors.append(f"{row_idx}행: {header} 필드는 필수입니다.")
                    break
                
                # 벤더 코드 정규화
                if field_name == "vendor" and value:
                    vendor_lower = value.lower().strip()
                    vendor_map = {
                        "palo alto": "paloalto",
                        "paloalto": "paloalto",
                        "ngf": "ngf",
                        "secui ngf": "ngf",
                        "mock": "mock"
                    }
                    value = vendor_map.get(vendor_lower, vendor_lower)
                    if value not in ["paloalto", "ngf", "mock"]:
                        errors.append(f"{row_idx}행: 잘못된 벤더 코드입니다. (paloalto, ngf, mock 중 하나)")
                        break
                
                # SSH 옵션 처리
                if field_name == "use_ssh_for_last_hit_date":
                    if isinstance(value, str):
                        value = value.upper() in ["TRUE", "1", "YES", "Y"]
                    elif value is None:
                        value = False
                    else:
                        value = bool(value)
                
                device_data[field_name] = value
            
            # 필수 필드가 모두 있는 경우에만 추가
            if all(device_data.get(field) for field in ["name", "ip_address", "vendor", "username", "password"]):
                devices.append((row_idx, device_data))
        
        if errors:
            raise HTTPException(status_code=400, detail="\n".join(errors))
        
        if not devices:
            raise HTTPException(status_code=400, detail="등록할 장비 데이터가 없습니다.")
        
        # 중복 확인 및 등록
        success_count = 0
        failed_devices = []
        
        for row_idx, device_data in devices:
            try:
                # 이름 중복 확인
                existing_device = await crud.device.get_device_by_name(db, name=device_data["name"])
                if existing_device:
                    failed_devices.append(f"{row_idx}행: 이름 '{device_data['name']}'이 이미 존재합니다.")
                    continue
                
                # IP 주소 중복 확인
                result = await db.execute(
                    select(models.Device).filter(models.Device.ip_address == device_data["ip_address"])
                )
                existing_ip = result.scalars().first()
                if existing_ip:
                    failed_devices.append(f"{row_idx}행: IP 주소 '{device_data['ip_address']}'이 이미 존재합니다.")
                    continue
                
                # 장비 생성
                device_create = schemas.DeviceCreate(
                    name=device_data["name"],
                    ip_address=device_data["ip_address"],
                    vendor=device_data["vendor"],
                    username=device_data["username"],
                    password=device_data["password"],
                    password_confirm=device_data["password"],
                    ha_peer_ip=device_data.get("ha_peer_ip"),
                    description=device_data.get("description"),
                    use_ssh_for_last_hit_date=device_data.get("use_ssh_for_last_hit_date", False)
                )
                
                await crud.device.create_device(db=db, device=device_create)
                success_count += 1
                
            except Exception as e:
                logger.error(f"Failed to import device at row {row_idx}: {e}", exc_info=True)
                failed_devices.append(f"{row_idx}행: {str(e)}")
        
        result_message = f"총 {len(devices)}개 중 {success_count}개 장비가 등록되었습니다."
        if failed_devices:
            result_message += f"\n\n실패한 항목:\n" + "\n".join(failed_devices)
        
        return {
            "success": True,
            "total": len(devices),
            "success_count": success_count,
            "failed_count": len(failed_devices),
            "message": result_message,
            "failed_devices": failed_devices
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to import devices from Excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"엑셀 파일 처리 실패: {str(e)}")
